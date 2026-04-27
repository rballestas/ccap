"""Input handling for CSV and Excel source files."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


class InputHandler:
    """Loads structured input files into normalized record dictionaries."""

    def __init__(self, supported_extensions: set[str]) -> None:
        self.supported_extensions = supported_extensions

    def load_data(self, file_path: Path) -> list[dict[str, Any]]:
        if not file_path.exists():
            raise FileNotFoundError(f"No se encontro el archivo de entrada: {file_path}")

        extension = file_path.suffix.lower()
        if extension not in self.supported_extensions:
            allowed = ", ".join(sorted(self.supported_extensions))
            raise ValueError(f"Formato de entrada no soportado: {extension}. Permitidos: {allowed}")

        if extension == ".csv":
            return self._load_csv(file_path)

        return self._load_excel(file_path)

    def _load_csv(self, file_path: Path) -> list[dict[str, Any]]:
        with file_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            return [self._normalize_record(record) for record in reader]

    def _load_excel(self, file_path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Para leer Excel instala la dependencia openpyxl.") from exc

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(header).strip() if header is not None else "" for header in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {headers[index]: self._cell_to_string(value) for index, value in enumerate(row) if index < len(headers) and headers[index]}
            records.append(self._normalize_record(record))
        return records

    def _normalize_record(self, record: dict[str, Any]) -> dict[str, Any]:
        return {str(key).strip(): self._cell_to_string(value) for key, value in record.items()}

    def _cell_to_string(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
